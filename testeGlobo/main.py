import openpyxl

# Carregar o arquivo Excel existente
arquivo = 'teste1.xlsx'
planilha = openpyxl.load_workbook(arquivo)
pagina_planilha = planilha['Plan8']

# Inicializar o total de horas
totalHorasNovela = 0
countPlayNovelaA =0
countPlayNovelaC=0
totalHorasSerie = 0
countPlaySerie =0
# Percorrer as linhas da planilha e calcular o total de horas para novelas
for linha in range(5, 30):  # Linhas de 5 a 29
    celulaIdConteudo = pagina_planilha.cell(row=linha, column=7)
    valorIdLinha = celulaIdConteudo.value

    # Verificar se o valorIdLinha é 10406 ou 10206
    if valorIdLinha == 10406 or valorIdLinha == 10206:
        if valorIdLinha == 10406:
            countPlayNovelaA += 1
        if valorIdLinha == 10206:
            countPlayNovelaC += 1
        valorHorasNovela = pagina_planilha.cell(row=linha, column=9).value
        totalHorasNovela += valorHorasNovela


    if valorIdLinha == 10352 or valorIdLinha == 10835:
        countPlaySerie += 1
        valorHorasSerie = pagina_planilha.cell(row=linha, column=9).value
        totalHorasSerie += valorHorasSerie




if countPlayNovelaA>countPlayNovelaC:
    primeiraRanking = "A"
    segundaRanking = "C"
elif countPlayNovelaC>countPlayNovelaA:
    primeiraRanking = "C"
    segundaRanking = "A"
#novela = 10406 e 10206
#serie = 10352 e 10835



totalHorasNovelaFormatado = f"{int(totalHorasNovela):02}:{int((totalHorasNovela * 60) % 60):02}"
totalHorasSerieFormatado = f"{int(totalHorasSerie):02}:{int((totalHorasSerie * 60) % 60):02}"

# Imprimir o valor total de horas consumidas em novelas
print(f"\nO valor total de horas consumidas em \033[1mNOVELAS\033[0m é de: \033[1m{totalHorasNovelaFormatado}h\033[0m com um total de \033[1m{countPlayNovelaA+countPlayNovelaC}\033[0m plays.")
print(f"O valor total de horas consumidas em \033[1mSÉRIES\033[0m é de: \033[1m{totalHorasSerieFormatado}h\033[0m com um total de \033[1m{countPlaySerie}\033[0m plays.")
print()
print(f"Segue o ranking das novelas mais assistidas:\n\033[1m1º Lugar -> Novela {primeiraRanking}.\n2º Lugar -> Novela {segundaRanking}.\033[0m")


























#imprimir dados de cada linha
#for rows in pagina_planilha.iter_rows(min_row=4, max_row=8):
 #   print(rows[1].value, "   ", rows[2].value, "   ", rows[3].value)

#print('---------------------------------------------')








#Aqui eu faço a mesclagem e a centralizacao dos valores de duas celulas
####pagina_planilha.merge_cells(start_row=17, start_column=M, end_row=17, end_column=N)

#celula_mesclada.alignment = Alignment(horizontal='center', vertical='center')


#celulaMergeHorasAssistidas1 = pagina_planilha.cell(row=17, column=13)
#celulaMergeHorasAssistidas2 = pagina_planilha.cell(row=17, column=14)

#pagina_planilha.merge_cells(celulaMergeHorasAssistidas1, celulaMergeHorasAssistidas2)

#pagina_planilha.cell(row=17, column=13, value="Quantidades de horas assistidas")


#teste = 'GLOBO!'
#celula = pagina_planilha.cell(row=20, column=12, value=teste)
#celula.value = teste
#planilha.cell(row=20, column=12, value=teste)




##planilha.save(arquivo)